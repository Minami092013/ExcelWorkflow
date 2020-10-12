#!/Library/Frameworks/Python.framework/Versions/3.8/bin/python3
#Check inventory status of parts in "Part List" and return \
#quantities needed for the number of builds specified by user

import openpyxl, collections

wb = openpyxl.load_workbook('/Users/Faye/Documents/coding_files/Inventory.xlsx')
stock = collections.defaultdict(int)
sheet = wb.get_sheet_by_name('Sheet1')

for i in range(2, sheet.max_row + 1):
	part = sheet.cell(row = i, column = 1).value
	stock[part] += int(sheet.cell(row = i, column = 5).value)

wb1 = openpyxl.load_workbook('/Users/Faye/Documents/coding_files/Part List.xlsx')
sheet1 = wb1.get_sheet_by_name('Sheet1')

number_of_builds = int(input("Enter Number of Builds Needed: "))

for i in range(2, sheet1.max_row + 1):
	part = sheet1.cell(row = i, column = 1).value
	
	qty_per_vehicle = int(sheet1.cell(row = i, column = 3).value)
	
	qty_to_purchase = qty_per_vehicle * number_of_builds - stock[part]
	
	if qty_to_purchase > 0:
		sheet1.cell(row = i, column = 4).value = qty_to_purchase
	else:
		sheet1.cell(row = i, column = 4).value = 0

wb1.save('/Users/Faye/Documents/coding_files/Part List_updated.xlsx')




